import openpyxl
import tkinter

# global variables (share)
excel_file_name = 'test2.xlsx'
workbook = openpyxl.load_workbook(excel_file_name)
sheet = workbook.active

def excel_config():
    # resize columns
    sheet.column_dimensions['A'].width = 20
    sheet.column_dimensions['B'].width = 20
    sheet.column_dimensions['C'].width = 20
    sheet.column_dimensions['D'].width = 20
    sheet.column_dimensions['E'].width = 20

    sheet.cell(row=1, column=1).value = "Name"
    sheet.cell(row=1, column=2).value = "Age"
    sheet.cell(row=1, column=3).value = "Course"
    sheet.cell(row=1, column=4).value = "Number"
    sheet.cell(row=1, column=5).value = "Address"

def focus_name(event):
    name_field.focus_set()

def focus_age(event):
    age_field.focus_set()

def focus_course(event):
    course_field.focus_set()

def focus_number(event):
    number_field.focus_set()

def focus_address(event):
    address_field.focus_set()

def clear():
    name_field.delete(0, tkinter.END)
    age_field.delete(0, tkinter.END)
    course_field.delete(0, tkinter.END)
    number_field.delete(0, tkinter.END)
    address_field.delete(0, tkinter.END)

def insert():
    if name_field.get() == "" and \
        age_field.get() == "" and \
        course_field.get() == "" and \
        number_field.get() == "" and \
        address_field.get() == "":
        print("Invalid input")
    
    else:
        cur_row = sheet.max_row
        cur_col = sheet.max_column

        sheet.cell(row = cur_row + 1, column = 1).value = name_field.get()
        sheet.cell(row = cur_row + 1, column = 2).value = age_field.get()
        sheet.cell(row = cur_row + 1, column = 3).value = course_field.get()
        sheet.cell(row = cur_row + 1, column = 4).value = number_field.get()
        sheet.cell(row = cur_row + 1, column = 5).value = address_field.get()

        workbook.save(excel_file_name)

        name_field.focus_set()
        clear()

if __name__ == "__main__":

    window = tkinter.Tk()
    window.configure(background='light blue')

    window.title("Book Keeping")
    window.geometry("500x400")

    excel_config()

    heading = tkinter.Label(window, text="Form", bg='light blue')
    name = tkinter.Label(window, text="Name:", bg='light blue')
    age = tkinter.Label(window, text="Age:", bg='light blue')
    course = tkinter.Label(window, text="Course:", bg='light blue')
    number = tkinter.Label(window, text="Number:", bg='light blue')
    address = tkinter.Label(window, text="Address:", bg='light blue')

    heading.grid(row=0, column=1)
    name.grid(row=1, column=0)
    age.grid(row=2, column=0)
    course.grid(row=3, column=0)
    number.grid(row=4, column=0)
    address.grid(row=5, column=0)

    name_field = tkinter.Entry(window)
    age_field = tkinter.Entry(window)
    course_field = tkinter.Entry(window)
    number_field = tkinter.Entry(window)
    address_field = tkinter.Entry(window)

    name_field.bind('<Return>', focus_name)
    age_field.bind('<Return>', focus_age)
    course_field.bind('<Return>', focus_course)
    number_field.bind('<Return>', focus_number)
    address_field.bind('<Return>', focus_address)

    name_field.grid(row=1, column=1, ipadx="100")
    age_field.grid(row=2, column=1, ipadx="100")
    course_field.grid(row=3, column=1, ipadx="100")
    number_field.grid(row=4, column=1, ipadx="100")
    address_field.grid(row=5, column=1, ipadx="100")

    excel_config()

    submit_button = tkinter.Button(window, text="Submit", fg='Black', bg='light yellow', command=insert)
    submit_button.grid(row=7, column=1)
    window.mainloop()